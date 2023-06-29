import subprocess
import os
from color_text import ColorText

try:
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    raise SystemExit(
        "Process terminated: openpyxl module not found. Please make sure it is installed."
    )


class UniversalUpdateFeature:
    destination_project_path = ""
    ssh_key_path = ""
    feature_name = ""
    work_book = None
    remote_repo_name = set()

    def __init__(
        self,
        destination_project_path="?",
        ssh_key_path="?",
        feature_name="1.0.0-example_feature",
    ):
        self.destination_project_path = destination_project_path
        self.ssh_key_path = ssh_key_path
        self.feature_name = feature_name

    def process(self):
        try:
            self.load_feature_excel()
            self.check_path()
            self.fetch_destination_origin()
            self.get_remote_repo_name()
            self.fetch_source_origin()
            self.create_new_branch_destination()
            self.process_cherry_pick()
            self.remove_remote_source()
            ColorText.print_ok_info(f"Perfect! Operation Completed!")
        except KeyboardInterrupt:
            self.remove_remote_source()
            raise SystemExit(ColorText.color_error("Process terminated by user"))

    def load_feature_excel(self):
        # load feature excel
        ColorText.print_ok_info(f"Loading Feature {self.feature_name}")
        try:
            # load work book at the destination project path
            self.work_book = load_workbook(
                f"{self.destination_project_path}/feature_xlsx/{self.feature_name}/{self.feature_name}.xlsx"
            )
        except:
            raise SystemExit(
                ColorText.color_error(
                    f"Feature excel file at Destination project path/feature_xlsx/{self.feature_name}/{self.feature_name}.xlsx not found."
                )
            )

    def check_path(self):
        ColorText.print_ok_info(
            f"Checking Destination Path {self.destination_project_path}"
        )
        # check if the path to folder exists
        if not os.path.exists(self.destination_project_path):
            raise SystemExit(ColorText.color_error("Path does not exist"))

        ColorText.print_ok_info(f"Checking SSH Key Path {self.ssh_key_path}")
        # check if the path to ssh key exists
        if not os.path.exists(self.ssh_key_path):
            raise SystemExit(ColorText.color_error("Path does not exist"))

    def fetch_destination_origin(self):
        ColorText.print_ok_info(f"Fetch Destination Origin")
        # fetch and create new branch on destination project
        try:
            subprocess.run(
                ["ssh-add", self.ssh_key_path],
                check=True,
                cwd=self.destination_project_path,
            )
            subprocess.run(
                ["git", "fetch", "--all"], check=True, cwd=self.destination_project_path
            )
            subprocess.run(
                ["git", "stash", "-u"], check=True, cwd=self.destination_project_path
            )
            subprocess.run(
                ["git", "branch", "-a"], check=True, cwd=self.destination_project_path
            )

            selected_branch = input(
                ColorText.color_warning("Input Select Branch (Enter for origin/main):")
            )
            if not selected_branch:
                selected_branch = "origin/main"

            subprocess.run(
                ["git", "checkout", selected_branch],
                check=True,
                cwd=self.destination_project_path,
            )
        except subprocess.CalledProcessError:
            raise SystemExit(ColorText.color_error("Fetch and Create branch failed"))

    def get_remote_repo_name(self):
        ColorText.print_ok_info(f"Get Remote Repo From {self.feature_name}.xlsx")
        # get remote repo name
        for row in self.work_book["Sheet1"]["B"]:
            if row.row == 1:
                continue
            self.remote_repo_name.add(row.value.strip())

        if len(self.remote_repo_name) == 0:
            raise SystemExit(
                ColorText.color_error(
                    f"No remote repo name found in {self.feature_name}.xlsx"
                )
            )

    def fetch_source_origin(self):
        try:
            current_origin = subprocess.run(
                ["git", "config", "--get", "remote.origin.url"],
                check=True,
                cwd=self.destination_project_path,
                capture_output=True,
                text=True,
            ).stdout.strip()

            # fetch and add remote for the source project
            for index, repo_name in enumerate(self.remote_repo_name):
                ColorText.print_ok_info(f"Fetch Remote Repo From {repo_name}")
                if repo_name == current_origin:
                    continue
                subprocess.run(
                    [
                        "git",
                        "remote",
                        "add",
                        f"source_project_{index}",
                        repo_name,
                    ],
                    check=True,
                    cwd=self.destination_project_path,
                )
                subprocess.run(
                    ["git", "fetch", f"source_project_{index}"],
                    check=True,
                    cwd=self.destination_project_path,
                )
        except subprocess.CalledProcessError:
            self.remove_remote_source()
            raise SystemExit(
                ColorText.color_error(f"Fetch failed, Repo name: {repo_name} Not Exist")
            )

    def create_new_branch_destination(self):
        ColorText.print_ok_info(f"Create New Branch {self.feature_name}")
        # create new branch
        try:
            subprocess.run(
                [
                    "git",
                    "switch",
                    "-c",
                    self.feature_name,
                ],
                check=True,
                cwd=self.destination_project_path,
            )
        except subprocess.CalledProcessError:
            self.remove_remote_source()
            raise SystemExit(ColorText.color_error("Create branch failed"))

    def process_cherry_pick(self):
        for index, row in enumerate(self.work_book["Sheet1"]):
            if index == 0:
                continue
            subprocess.run(
                ["git", "cherry-pick", row[2].value], cwd=self.destination_project_path
            )
            ColorText.print_ok_info(f"Cherry Pick {row[2].value} Success Please Check for Conflict")
            log = ""
            while not log:
                log = input(
                    ColorText.color_warning(
                        f"Input Code For Commit {row[2].value} [Complete[c]/Error[e]/Nothing[n]]:"
                    )
                )
                if log not in ["c", "e", "n"]:
                    log = ""
            status_cell = self.work_book["Sheet1"].cell(row=row[2].row, column=4)
            if log == "c":
                status_cell.value = "Complete"
                fill_cell = PatternFill(patternType="solid", fgColor="2E75B5")
                status_cell.fill = fill_cell
            elif log == "e":
                status_cell.value = "Error"
                fill_cell = PatternFill(patternType="solid", fgColor="C65911")
                status_cell.fill = fill_cell
            elif log == "n":
                status_cell.value = "Nothing"
                fill_cell = PatternFill(patternType="solid", fgColor="262626")
                status_cell.fill = fill_cell
            # save the workbook at the same place
            self.work_book.save(
                f"{self.destination_project_path}/feature_xlsx/{self.feature_name}/{self.feature_name}.xlsx"
            )

    def remove_remote_source(self):
        # remove remote source
        remote_repo = subprocess.run(
            ["git", "remote"],
            check=True,
            cwd=self.destination_project_path,
            capture_output=True,
            text=True,
        )
        existing_remotes = remote_repo.stdout.strip().split("\n")
        for remote_name in existing_remotes:
            if remote_name != "origin":
                try:
                    ColorText.print_ok_info(f"Remove Remote Source {remote_name}")
                    subprocess.run(
                        [
                            "git",
                            "remote",
                            "remove",
                            remote_name,
                        ],
                        check=True,
                        cwd=self.destination_project_path,
                    )
                except subprocess.CalledProcessError:
                    raise SystemExit(
                        ColorText.color_error(f"Remove remote {remote_name} failed")
                    )
