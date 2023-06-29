from openpyxl.styles import PatternFill


def adding_status_for_result_excel(log):
    if log == "c":
        fill_cell = PatternFill(patternType="solid", fgColor="2E75B5")
        return "Complete", fill_cell
    elif log == "e":
        fill_cell = PatternFill(patternType="solid", fgColor="C65911")
        return "Error", fill_cell
    elif log == "n":
        fill_cell = PatternFill(patternType="solid", fgColor="262626")
        return "Nothing", fill_cell
